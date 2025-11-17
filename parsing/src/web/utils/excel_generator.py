from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import pandas as pd

class ExcelGenerator:
    """Класс для генерации Excel файлов"""
    
    def __init__(self):
        pass
    
    def generate_excel(self, data, fields: List[str], output_path: str):
        """
        Генерация Excel файла из обработанных данных
        
        Args:
            data: Список словарей с обработанными данными
            fields: Список полей для включения в Excel
            output_path: Путь для сохранения файла
        """

        screens, synopses = data

        df = pd.DataFrame({'id': [s['id'] for s in screens],
        'text': [s['text'] for s in screens],
        'title': [s['title'] for s in screens], 
        'Синопсис': synopses})

        df.to_excel(output_path, index = False)

        # wb = Workbook()
        # ws = wb.active
        # ws.title = "Обработанные данные"
        
        # # Стили для заголовков
        # header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # header_font = Font(bold=True, color="FFFFFF", size=12)
        # header_alignment = Alignment(horizontal="center", vertical="center")
        # border = Border(
        #     left=Side(style='thin'),
        #     right=Side(style='thin'),
        #     top=Side(style='thin'),
        #     bottom=Side(style='thin')
        # )
        
        # # Записываем заголовки
        # for col_idx, field in enumerate(fields, start=1):
        #     cell = ws.cell(row=1, column=col_idx, value=field)
        #     cell.fill = header_fill
        #     cell.font = header_font
        #     cell.alignment = header_alignment
        #     cell.border = border
        
        # # Записываем данные
        # for row_idx, record in enumerate(data, start=2):
        #     for col_idx, field in enumerate(fields, start=1):
        #         value = record.get(field, '')
        #         cell = ws.cell(row=row_idx, column=col_idx, value=value)
        #         cell.border = border
        #         cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # # Автоматическая ширина колонок
        # for col_idx, field in enumerate(fields, start=1):
        #     column_letter = get_column_letter(col_idx)
        #     max_length = 0
            
        #     # Проверяем длину заголовка
        #     max_length = max(max_length, len(str(field)))
            
        #     # Проверяем длину данных в колонке
        #     for row in ws[column_letter]:
        #         try:
        #             if row.value:
        #                 max_length = max(max_length, len(str(row.value)))
        #         except:
        #             pass
            
        #     # Устанавливаем ширину с небольшим запасом
        #     adjusted_width = min(max_length + 2, 50)
        #     ws.column_dimensions[column_letter].width = adjusted_width
        
        # # Фиксируем первую строку (заголовки)
        # ws.freeze_panes = 'A2'
        
        # # Сохраняем файл
        # wb.save(output_path)


